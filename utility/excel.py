from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
import numpy as np
import sqlite3
from utility import helpers

positions = ['QB', 'RB', 'WR', 'TE']


def save_to_excel(position_dfs, filename='fantasy_football_auction_values.xlsx'):
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    for pos in positions:
        ws = wb.create_sheet(pos)
        df = position_dfs[pos][['Name', 'Team', 'ModelPoints', 'VORP', 'AuctionValue']]

        headers = ['Rank'] + list(df.columns)
        ws.append(headers)

        for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            ws.append([i - 1] + list(row))

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        for column in ws.columns:
            max_length = 0
            column_cells = [cell for cell in column]
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    ws_summary = wb.create_sheet('Summary', 0)
    summary_data = []
    for pos in positions:
        top_players = position_dfs[pos].head(10)
        summary_data.extend(
            top_players[['Name', 'Position', 'Team', 'ModelPoints', 'VORP', 'AuctionValue']].values.tolist()
        )

    summary_df = pd.DataFrame(
        summary_data,
        columns=['Name', 'Position', 'Team', 'ModelPoints', 'VORP', 'AuctionValue']
    )
    summary_df = summary_df.sort_values('AuctionValue', ascending=False)

    ws_summary.append(['Rank'] + list(summary_df.columns))
    for i, row in enumerate(dataframe_to_rows(summary_df, index=False, header=False), start=2):
        ws_summary.append([i - 1] + list(row))

    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for row in ws_summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    for column in ws_summary.columns:
        max_length = 0
        column_cells = [cell for cell in column]
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        ws_summary.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    wb.save(filename)
    print(f"Excel file '{filename}' has been created successfully.")


def export_draft_history(filename='draft_history.xlsx'):
    conn = sqlite3.connect(helpers.DB_FILE)
    df = pd.read_sql_query('SELECT * FROM draft_picks', conn)
    conn.close()
    df.to_excel(filename, index=False)
    print(f"Draft history exported to {filename}")
